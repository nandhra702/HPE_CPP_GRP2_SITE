from django.contrib.admin.views.decorators import staff_member_required
from django.core.paginator import Paginator
from django.db.models import Q
from django.http import JsonResponse
from django.shortcuts import render
from django.views.decorators.http import require_GET

from judge.models import Problem, MCQQuestion, ProblemType, ProblemGroup

@staff_member_required
@require_GET
def contest_dashboard(request):
    return render(request, 'admin/judge/contest/dashboard.html')

@staff_member_required
@require_GET
def contest_dashboard_api(request):
    # Parameters
    page = int(request.GET.get('page', 1))
    limit = int(request.GET.get('limit', 25))
    search = request.GET.get('search', '')
    category = request.GET.get('category', '') # Easy, Medium, Hard (mapped to group or type)
    p_type = request.GET.get('type', '')
    p_type = request.GET.get('type', '')
    is_mcq = request.GET.get('is_mcq', 'false') == 'true'
    ids_filter = request.GET.get('ids', '')
    
    if is_mcq:
        queryset = MCQQuestion.objects.all()
        
        # Filter by specific IDs if provided
        if ids_filter:
            try:
                ids_list = [int(x) for x in ids_filter.split(',') if x.strip()]
                queryset = queryset.filter(id__in=ids_list)
            except ValueError:
                pass

        if search:
            queryset = queryset.filter(
                Q(code__icontains=search) | 
                Q(description__icontains=search)
            )
        if category:
            # Try matching group or type
            queryset = queryset.filter(
                Q(group__name__iexact=category) | 
                Q(group__full_name__iexact=category) |
                Q(types__name__iexact=category) |
                Q(types__full_name__iexact=category)
            ).distinct()
        
        if p_type:
             queryset = queryset.filter(types__id=p_type)

        # Additional MCQ filters
        mcq_format = request.GET.get('format', '')
        if mcq_format:
            queryset = queryset.filter(question_type=mcq_format)

        paginator = Paginator(queryset.order_by('code'), limit)
        page_obj = paginator.get_page(page)
        
        data = {
            'items': [{
                'id': q.id,
                'code': q.code,
                'name': q.description[:50] + '...' if len(q.description) > 50 else q.description,
                'points': q.points,
                'type': q.question_type,
                'group': q.group.full_name if q.group else 'Uncategorized',
                'types': [t.name for t in q.types.all()]
            } for q in page_obj],
            'total': paginator.count,
            'num_pages': paginator.num_pages,
            'current_page': page
        }
    else:
        queryset = Problem.objects.all()
        
        # Filter by specific IDs if provided
        if ids_filter:
            try:
                ids_list = [int(x) for x in ids_filter.split(',') if x.strip()]
                queryset = queryset.filter(id__in=ids_list)
            except ValueError:
                pass
                
        if search:
            queryset = queryset.filter(
                Q(code__icontains=search) | 
                Q(name__icontains=search)
            )
        
        if category:
            # User requested filtering by "Easy", "Medium", "Hard"
            # We check if they match a group or type name
            queryset = queryset.filter(
                Q(group__name__iexact=category) | 
                Q(group__full_name__iexact=category) |
                Q(types__name__iexact=category) |
                Q(types__full_name__iexact=category)
            ).distinct()

        if p_type:
            queryset = queryset.filter(types__id=p_type)

        paginator = Paginator(queryset.order_by('code'), limit)
        page_obj = paginator.get_page(page)

        data = {
            'items': [{
                'id': p.id,
                'code': p.code,
                'name': p.name,
                'points': p.points,
                'group': p.group.full_name if p.group else 'Uncategorized',
                'types': [t.name for t in p.types.all()]
            } for p in page_obj],
            'total': paginator.count,
            'num_pages': paginator.num_pages,
            'current_page': page
        }

    return JsonResponse(data)

@staff_member_required
@require_GET
def contest_dashboard_metadata(request):
    # Return available types and groups for filters
    types = ProblemType.objects.values('id', 'name', 'full_name')
    groups = ProblemGroup.objects.values('id', 'name', 'full_name')
    return JsonResponse({
        'types': list(types),
        'groups': list(groups)
    })


from django.views.decorators.http import require_POST
from django.views.decorators.csrf import csrf_exempt
from django.utils.text import slugify
from django.utils import timezone
import csv
import io

@staff_member_required
@require_POST
def contest_dashboard_upload(request):
    """Handle CSV upload for Problems or MCQs from the contest dashboard."""
    upload_type = request.POST.get('type', 'problem')  # 'problem' or 'mcq'
    csv_file = request.FILES.get('file')
    
    if not csv_file:
        return JsonResponse({'error': 'No file uploaded'}, status=400)
    
    try:
        file_name = csv_file.name.lower()
        rows = []
        
        # Handle Excel files (.xlsx)
        if file_name.endswith('.xlsx'):
            from openpyxl import load_workbook
            wb = load_workbook(filename=io.BytesIO(csv_file.read()))
            ws = wb.active
            for row in ws.iter_rows(values_only=True):
                rows.append([str(cell) if cell is not None else '' for cell in row])
        else:
            # Handle CSV files
            decoded_file = csv_file.read().decode('utf-8').splitlines()
            reader = csv.reader(decoded_file)
            rows = list(reader)
        
        if upload_type == 'mcq':
            return _process_mcq_upload(rows, request)
        else:
            return _process_problem_upload(rows, request)
            
    except Exception as e:
        return JsonResponse({'error': f'Error processing file: {str(e)}'}, status=500)


def _process_problem_upload(rows, request):
    """Process Problem CSV upload with test case file creation."""
    from judge.models import Problem, Language, Solution
    from judge.models.problem import ProblemGroup
    from django.contrib.auth.models import User
    import os
    import io
    import zipfile
    from django.conf import settings
    
    # Get or create difficulty groups dynamically (matching HPE bulk upload)
    easy_group, _ = ProblemGroup.objects.get_or_create(
        name='easy', defaults={'full_name': 'Easy'}
    )
    medium_group, _ = ProblemGroup.objects.get_or_create(
        name='medium', defaults={'full_name': 'Medium'}
    )
    hard_group, _ = ProblemGroup.objects.get_or_create(
        name='hard', defaults={'full_name': 'Hard'}
    )
    
    difficulty_map = {
        'easy': easy_group.id,
        'medium': medium_group.id,
        'hard': hard_group.id,
    }
    
    # Get all admin users as authors
    admin_profiles = []
    admin_users = User.objects.filter(is_staff=True, is_active=True)
    for user in admin_users:
        if hasattr(user, 'profile'):
            admin_profiles.append(user.profile)
    
    created = []
    duplicates = []
    errors = []
    
    for row_num, row in enumerate(rows, 1):
        if len(row) < 6:
            continue
        if all(not cell or str(cell).strip() == '' for cell in row):
            continue
        
        name = row[0].strip()
        if not name:
            continue
        
        body = row[1]
        constraints = row[2]
        
        # Time limit
        tl_value = row[3].strip().lower() if row[3] else ''
        time_limit = 60.0
        if tl_value and tl_value not in ('null', 'none', ''):
            try:
                time_limit = float(row[3])
            except:
                pass
        
        # Memory limit
        ml_value = row[4].strip().lower() if row[4] else ''
        memory_limit = 524288
        if ml_value and ml_value not in ('null', 'none', ''):
            try:
                memory_limit = int(row[4])
            except:
                pass
        
        # Difficulty
        raw_difficulty = row[5] if len(row) > 5 and row[5] else ''
        difficulty_value = raw_difficulty.strip().lower()
        
        if difficulty_value not in difficulty_map:
            errors.append({'row': row_num, 'name': name, 'error': f"Invalid difficulty '{raw_difficulty}'. Must be easy, medium, or hard."})
            continue
        group_id = difficulty_map[difficulty_value]
        
        # Solution (column 6 - optional)
        solution_content = row[6].strip() if len(row) > 6 and row[6] else ''
        
        # Test cases start at column 7 (after solution)
        test_cases_data = row[7:] if len(row) > 7 else []
        
        # Check for duplicate
        existing = Problem.objects.filter(name=name).first()
        if existing:
            duplicates.append({
                'row': row_num,
                'name': name,
                'message': 'Problem with this name already exists.',
                'id': existing.id,
                'code': existing.code,
                'points': existing.points,
                'group': existing.group.full_name if existing.group else 'Uncategorized'
            })
            continue
        
        # Generate unique code
        base_code = slugify(name)[:20] or "prob"
        code = base_code
        suffix = 1
        while Problem.objects.filter(code=code).exists():
            code = f"{base_code}{suffix}"
            suffix += 1
        
        # Build description with examples (first 2 test cases)
        examples_md = ""
        example_num = 1
        
        for i in range(0, len(test_cases_data), 2):
            if i + 1 >= len(test_cases_data):
                break
            in_data = test_cases_data[i]
            out_data = test_cases_data[i+1]
            if example_num <= 2:
                examples_md += f"\n## Example {example_num}\n**Input:**\n```\n{in_data}\n```\n\n**Output:**\n```\n{out_data}\n```\n"
            example_num += 1
        
        description = body
        if examples_md:
            description += f"\n{examples_md}"
        if constraints and constraints.lower() != 'none':
            description += f"\n## Constraints\n{constraints}\n"
        
        # Create problem
        try:
            problem = Problem.objects.create(
                code=code,
                name=name,
                description=description,
                time_limit=time_limit,
                memory_limit=memory_limit,
                points=0,
                is_public=True,
                is_manually_managed=True,
                group_id=group_id,
                date=timezone.now()
            )
            
            # Add ALL languages (per user request)
            all_languages = Language.objects.all()
            if all_languages.exists():
                problem.allowed_languages.set(all_languages)
            
            # Add admin users as authors
            if admin_profiles:
                problem.authors.set(admin_profiles)
            
            # Create Solution if provided
            if solution_content:
                Solution.objects.create(
                    problem=problem,
                    content=solution_content,
                    is_public=False,
                    publish_on=timezone.now()
                )
            
            # Create test case files (matching HPE bulk upload structure)
            if test_cases_data:
                problems_dir = os.path.join(settings.BASE_DIR, 'problems')
                problem_dir = os.path.join(problems_dir, code)
                os.makedirs(problem_dir, exist_ok=True)
                
                # Build init.yml content
                init_yml_cases = []
                zip_buffer = io.BytesIO()
                
                with zipfile.ZipFile(zip_buffer, 'w') as zf:
                    case_idx = 1
                    for i in range(0, len(test_cases_data), 2):
                        if i + 1 >= len(test_cases_data):
                            break
                        
                        in_data = test_cases_data[i]
                        out_data = test_cases_data[i+1]
                        
                        in_name = f"{case_idx}.in"
                        out_name = f"{case_idx}.out"
                        
                        # Write individual files to folder
                        with open(os.path.join(problem_dir, in_name), 'w') as f:
                            f.write(str(in_data))
                        with open(os.path.join(problem_dir, out_name), 'w') as f:
                            f.write(str(out_data))
                        
                        # Add to zip
                        zf.writestr(in_name, str(in_data))
                        zf.writestr(out_name, str(out_data))
                        
                        init_yml_cases.append({'in': in_name, 'out': out_name})
                        case_idx += 1
                
                if init_yml_cases:
                    # Write init.yml
                    yml_lines = [
                        f"name: {name}",
                        f"code: {code}",
                        "type: standard",
                        "validator: token",
                        "limits:",
                        f"  time: {time_limit}",
                        f"  memory: {memory_limit}",
                        "archive: testcases.zip",
                        "cases:",
                    ]
                    for case in init_yml_cases:
                        yml_lines.append(f"  - in: {case['in']}")
                        yml_lines.append(f"    out: {case['out']}")
                    
                    with open(os.path.join(problem_dir, 'init.yml'), 'w') as f:
                        f.write('\n'.join(yml_lines) + '\n')
                    
                    # Write testcases.zip
                    with open(os.path.join(problem_dir, 'testcases.zip'), 'wb') as f:
                        f.write(zip_buffer.getvalue())
            
            created.append({
                'id': problem.id,
                'code': problem.code,
                'name': problem.name,
                'points': problem.points,
                'group': problem.group.full_name if problem.group else 'Uncategorized'
            })
        except Exception as e:
            errors.append({'row': row_num, 'name': name, 'error': str(e)})
    
    return JsonResponse({
        'created': created,
        'duplicates': duplicates,
        'errors': errors,
        'type': 'problem'
    })


def _process_mcq_upload(rows, request):
    """Process MCQ CSV upload."""
    from judge.models import MCQQuestion, MCQOption
    
    created = []
    duplicates = []
    errors = []
    
    for row_num, row in enumerate(rows, 1):
        if len(row) < 6:
            continue
        if all(not cell or str(cell).strip() == '' for cell in row):
            continue
        
        question_text = row[0].strip()
        if not question_text:
            continue
        
        # Get 4 options
        options = [row[i].strip() if i < len(row) else '' for i in range(1, 5)]
        
        # Get answer columns (column 6 onwards)
        answers_raw = [row[i].strip().lower() for i in range(5, len(row)) if i < len(row) and row[i] and row[i].strip()]
        
        # Match answers to options by text
        correct_options = set()
        for idx, option in enumerate(options):
            if option.lower() in answers_raw:
                correct_options.add(idx + 1)
        
        if not correct_options:
            errors.append({'row': row_num, 'name': question_text[:50], 'error': 'No matching answers found. Check that answer text matches option text exactly.'})
            continue
        
        question_type = 'MULTIPLE' if len(correct_options) > 1 else 'SINGLE'
        
        # Check for duplicate
        existing = MCQQuestion.objects.filter(description__startswith=question_text[:100]).first()
        if existing:
            duplicates.append({
                'row': row_num,
                'name': question_text[:50],
                'message': 'MCQ with similar text already exists.',
                'id': existing.id,
                'code': existing.code,
                'points': existing.points,
                'group': existing.group.full_name if existing.group else 'Uncategorized'
            })
            continue
        
        # Generate unique code
        base_code = slugify(question_text[:15]).replace('-', '') or "mcq"
        code = base_code[:20]
        suffix = 1
        while MCQQuestion.objects.filter(code=code).exists():
            code = f"{base_code[:17]}{suffix}"
            suffix += 1
        
        try:
            mcq = MCQQuestion.objects.create(
                code=code,
                description=question_text,
                question_type=question_type,
                points=1.0,
                is_public=True
            )
            
            # Create options
            for idx, option_text in enumerate(options):
                if option_text:
                    MCQOption.objects.create(
                        question=mcq,
                        option_text=option_text,
                        is_correct=(idx + 1) in correct_options,
                        order=idx
                    )
            
            created.append({
                'id': mcq.id,
                'code': mcq.code,
                'name': mcq.description[:50] + '...' if len(mcq.description) > 50 else mcq.description,
                'points': mcq.points,
                'group': mcq.group.full_name if mcq.group else 'Uncategorized'
            })
        except Exception as e:
            errors.append({'row': row_num, 'name': question_text[:50], 'error': str(e)})
    
    return JsonResponse({
        'created': created,
        'duplicates': duplicates,
        'errors': errors,
        'type': 'mcq'
    })

