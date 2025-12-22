from django.contrib import admin
from django.utils.translation import gettext_lazy as _
from django.shortcuts import get_object_or_404, render, redirect
from django.urls import path, reverse
from django.utils.html import format_html
from django.contrib import messages
from django.http import HttpResponseRedirect
from django.core.mail import send_mail
from django.conf import settings
from django.utils.crypto import get_random_string
from django.contrib.auth.models import User
from django.utils.decorators import method_decorator
from django.views.decorators.http import require_POST
from django.db import transaction
import json

from .sites import HPEAdminSite
from .forms import ContestParticipantUploadForm, HPEContestForm, HPEProblemForm, HPEMCQForm, HPEProblemBulkUploadForm, HPEMCQBulkUploadForm

from judge.models import Problem, MCQQuestion, Contest, ContestProblem, ContestMCQ, Profile, Solution
from judge.admin.problem import ProblemAdmin
from judge.admin.mcq import MCQQuestionAdmin
from judge.admin.contest import ContestAdmin, ContestForm, ContestProblemInline, ContestMCQInline

# Initialize the custom admin site
hpe_admin_site = HPEAdminSite(name='hpe_admin')

class HPEContestAdmin(ContestAdmin):
    form = HPEContestForm
    change_form_template = 'hpe_admin/contest_change_form.html'
    inlines = []  # Removed - using Problem Dashboard instead
    
    def get_form(self, request, obj=None, **kwargs):
        # We skip ContestAdmin.get_form because it assumes fields exist that we removed.
        # super(ContestAdmin, self) will resolve to NoBatchDeleteMixin -> SortableAdminBase -> VersionAdmin -> ModelAdmin
        form = super(ContestAdmin, self).get_form(request, obj, **kwargs)
        return form

    # Custom fieldsets as requested - Problems Dashboard at bottom, just above inlines
    fieldsets = (
        (None, {'fields': ('key', 'name', 'description')}),
        (_('Access Control'), {'fields': ('authors', 'testers')}),
        (_('Tester Permissions'), {'fields': ('tester_see_submissions', 'tester_see_scoreboard')}),
        (_('Scheduling'), {'fields': ('start_time', 'end_time', 'time_limit')}),
        (_('Participants'), {'fields': ('private_contestants', 'participants_csv')}),
        (_('Problems Dashboard'), {'fields': ('dashboard_button', 'contest_problems_json', 'contest_mcqs_json', 'contest_randomization_json')}),
    )

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path('<int:contest_id>/upload-participants/', self.upload_participants_view, name='hpe_contest_upload_participants'),
            path('<int:contest_id>/send-invites/', self.send_invites_view, name='hpe_contest_send_invites'),
        ]
        return custom_urls + urls

    def save_related(self, request, form, formsets, change):
        # 1. Save Inlines (Standard Django behavior)
        admin.ModelAdmin.save_related(self, request, form, formsets, change)
        
        # 2. Rescore logic
        if not self._rescored and any(formset.has_changed() for formset in formsets):
            self._rescore(form.cleaned_data['key'])

        # 3. JSON Logic (Non-destructive / Merging)
        if 'contest_problems_json' in form.cleaned_data and form.cleaned_data['contest_problems_json']:
            try:
                problems_data = json.loads(form.cleaned_data['contest_problems_json'])
                normalized_problems = []
                for item in problems_data:
                    if isinstance(item, int): normalized_problems.append({'id': item})
                    else: normalized_problems.append(item)
                
                current_problems = {cp.problem_id: cp for cp in form.instance.contest_problems.all()}
                
                for i, p_item in enumerate(normalized_problems):
                    pid = int(p_item['id'])
                    points = p_item.get('points')
                    partial = p_item.get('partial', True)
                    is_pretested = p_item.get('is_pretested', False)
                    max_submissions = p_item.get('max_submissions')
                    output_prefix_override = p_item.get('output_prefix_override', 0)
                    
                    if pid in current_problems:
                        cp = current_problems[pid]
                        changed = False
                        if cp.order != i: cp.order = i; changed = True
                        if points is not None and cp.points != points: cp.points = points; changed = True
                        if changed: cp.save()
                    else:
                        prob = Problem.objects.get(id=pid)
                        ContestProblem.objects.create(
                            contest=form.instance,
                            problem=prob,
                            points=points if points is not None else prob.points,
                            partial=partial,
                            is_pretested=is_pretested,
                            max_submissions=max_submissions,
                            output_prefix_override=output_prefix_override,
                            order=i
                        )
            except Exception as e: pass

        if 'contest_mcqs_json' in form.cleaned_data and form.cleaned_data['contest_mcqs_json']:
            try:
                mcq_data = json.loads(form.cleaned_data['contest_mcqs_json'])
                normalized_mcqs = []
                for item in mcq_data:
                    if isinstance(item, int): normalized_mcqs.append({'id': item})
                    else: normalized_mcqs.append(item)

                current_mcqs = {cm.mcq_question_id: cm for cm in form.instance.contest_mcqs.all()}
                
                for i, m_item in enumerate(normalized_mcqs):
                    mid = int(m_item['id'])
                    points = m_item.get('points')
                    if mid in current_mcqs:
                        cm = current_mcqs[mid]
                        changed = False
                        if cm.order != i: cm.order = i; changed = True
                        if points is not None and cm.points != points: cm.points = points; changed = True
                        if changed: cm.save()
                    else:
                        mcq = MCQQuestion.objects.get(id=mid)
                        ContestMCQ.objects.create(
                            contest=form.instance,
                            mcq_question=mcq,
                            points=points if points is not None else mcq.points,
                            order=i
                        )
            except Exception as e: pass

        # 4. JSON Logic for Randomization
        if 'contest_randomization_json' in form.cleaned_data and form.cleaned_data['contest_randomization_json']:
            try:
                rand_data = json.loads(form.cleaned_data['contest_randomization_json'])
                # rand_data structure: {'enabled': bool, 'config': {...}, 'regular_enabled': bool, 'mcq_enabled': bool}
                
                # Update the contest instance
                contest = form.instance
                contest.randomization_config = rand_data
                
                # Determine if randomization is active (legacy support + specific flags)
                is_enabled = rand_data.get('enabled', False)
                if rand_data.get('regular_enabled') or rand_data.get('mcq_enabled'):
                    is_enabled = True
                    
                contest.randomize = is_enabled
                contest.save()
                
            except Exception as e:
                # Log silently or maybe message user if critical
                print(f"Error saving randomization config: {e}")
                pass

        # 4. Handle Participant CSV Upload (Moved from save_model to ensure M2M persistence)
        if 'participants_csv' in request.FILES:
            csv_file = request.FILES['participants_csv']
            try:
                decoded_file = csv_file.read().decode('utf-8').splitlines()
                import csv
                reader = csv.reader(decoded_file)
                
                created_count = 0
                added_count = 0
                
                for row in reader:
                    if not row: continue
                    email = row[0].strip()
                    if not email or '@' not in email: continue
                    
                    email_val = row[0].strip()
                    if not email_val or '@' not in email_val: continue
                    
                    if len(row) > 1 and row[1].strip():
                         username = row[1].strip()
                    else:
                         username = email_val.split('@')[0]
                    
                    user, created = User.objects.get_or_create(
                        email=email_val,
                        defaults={'username': username}
                    )
                    
                    if created:
                        password = get_random_string(12)
                        user.set_password(password)
                        user.save()
                        Profile.objects.get_or_create(user=user)
                        created_count += 1
                    
                    profile = user.profile
                    # Use form.instance (obj)
                    if not form.instance.private_contestants.filter(id=profile.id).exists():
                        form.instance.private_contestants.add(profile)
                        added_count += 1
                
                if not form.instance.is_private and added_count > 0:
                    form.instance.is_private = True
                    form.instance.save()

                if added_count > 0:
                    self.message_user(request, f"Processed CSV: {created_count} accounts created, {added_count} participants added.", level=messages.SUCCESS)
                    
            except Exception as e:
                self.message_user(request, f"Error processing CSV: {e}", level=messages.ERROR)


    def upload_participants_view(self, request, contest_id):
        # ... kept for backward compatibility if needed, but primary method is now valid on save_model
        return redirect('hpe_admin:judge_contest_change', contest_id)

    def send_invites_view(self, request, contest_id):
        from django.http import JsonResponse
        
        contest = get_object_or_404(Contest, id=contest_id)
        
        # Iterate private contestants
        success_emails = []
        failed_emails = []
        skipped_emails = []
        
        for profile in contest.private_contestants.all():
            user = profile.user
            if not user.email:
                skipped_emails.append({
                    'username': user.username,
                    'reason': 'No email address configured'
                })
                continue
            
            # No password generation - using Google OAuth instead
            
            # Send Email with Google Sign-In instructions
            contest_url = request.build_absolute_uri(reverse('hpe_contest_landing', args=[contest.key]))
            subject = f"Invitation to {contest.name}"
            message = f"""
Hello {user.username},

You have been invited to participate in the contest "{contest.name}".

To access the contest, please visit the link below and sign in with your Google account:

Contest Link: {contest_url}

Make sure to use your Google account associated with this email address ({user.email}).

Good luck!
            """
            try:
                send_mail(subject, message, settings.DEFAULT_FROM_EMAIL, [user.email], fail_silently=False)
                success_emails.append(user.email)
            except Exception as e:
                failed_emails.append({
                    'email': user.email,
                    'reason': str(e)
                })
        
        # Check if this is an AJAX request
        is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
        
        if is_ajax:
            return JsonResponse({
                'success': True,
                'success_emails': success_emails,
                'failed_emails': failed_emails,
                'skipped_emails': skipped_emails,
                'total_sent': len(success_emails),
                'total_failed': len(failed_emails),
                'total_skipped': len(skipped_emails),
            })
        
        # Fallback for non-AJAX requests
        messages.success(request, f"Sent invites to {len(success_emails)} participants.")
        if failed_emails:
            failed_list = "<br>".join([f"{e['email']}: {e['reason']}" for e in failed_emails])
            messages.warning(request, f"Failed to send to {len(failed_emails)} participants: <br>" + failed_list)
            
        return redirect('hpe_admin:judge_contest_change', contest_id)



class HPEProblemAdmin(ProblemAdmin):
    form = HPEProblemForm
    change_list_template = 'hpe_admin/problem_change_list.html'
    change_form_template = 'hpe_admin/problem_change_form.html'
    
    # Show difficulty in list view (renamed from group)
    list_display = ['code', 'name', 'difficulty_display', 'points', 'is_public', 'date']
    list_filter = ['group', 'is_public']  # Filter by difficulty level
    search_fields = ['code', 'name']
    
    # Custom fieldsets with Difficulty dropdown
    fieldsets = (
        (None, {'fields': ('code', 'name', 'authors', 'testers')}),
        (_('Problem Body'), {'fields': ('description',)}),
        (_('Difficulty'), {'fields': ('group',)}),  # Difficulty dropdown
        (_('Resources'), {'fields': ('time_limit', 'memory_limit', 'points', 'partial')}),
        (_('Languages'), {'fields': ('allowed_languages',)}),
    )
    
    def difficulty_display(self, obj):
        """Display group as 'Difficulty' with colored styling"""
        if obj.group:
            return obj.group.full_name
        return '-'
    difficulty_display.short_description = 'Difficulty'
    difficulty_display.admin_order_field = 'group'
    
    def get_actions(self, request):
        """Re-enable delete action with custom implementation to handle distinct()"""
        actions = super().get_actions(request)
        # Add custom delete action that works with distinct querysets
        actions['delete_selected'] = (self.delete_selected_problems, 'delete_selected', _('Delete selected problems'))
        return actions
    
    def delete_selected_problems(self, modeladmin, request, queryset):
        """Custom delete action with confirmation page that handles distinct querysets"""
        from django.contrib import messages
        from django.template.response import TemplateResponse
        
        # Get the IDs first to avoid the distinct() issue
        problem_ids = list(queryset.values_list('id', flat=True))
        problems_to_delete = Problem.objects.filter(id__in=problem_ids)
        count = len(problem_ids)
        
        if count == 0:
            messages.warning(request, "No problems selected.")
            return None
        
        # Check if user confirmed deletion
        if request.POST.get('post') == 'yes':
            import os
            import shutil
            from django.conf import settings
            
            # Get problem codes before deletion (for folder cleanup)
            problem_codes = list(problems_to_delete.values_list('code', flat=True))
            
            # Delete from database
            problems_to_delete.delete()
            
            # Delete problem folders from filesystem
            problems_dir = os.path.join(settings.BASE_DIR, 'problems')
            deleted_folders = 0
            for code in problem_codes:
                folder_path = os.path.join(problems_dir, code)
                if os.path.exists(folder_path):
                    try:
                        shutil.rmtree(folder_path)
                        deleted_folders += 1
                    except Exception as e:
                        messages.warning(request, f"Could not delete folder for '{code}': {e}")
            
            messages.success(request, f"Successfully deleted {count} problem{'' if count == 1 else 's'} and {deleted_folders} folder{'' if deleted_folders == 1 else 's'}.")
            return None
        
        # Show confirmation page
        context = {
            **self.admin_site.each_context(request),
            'title': _('Are you sure?'),
            'problems': problems_to_delete,
            'count': count,
            'action_checkbox_name': '_selected_action',
            'queryset': queryset,
            'opts': self.model._meta,
            'media': self.media,
        }
        
        return TemplateResponse(request, 'hpe_admin/confirm_delete_problems.html', context)
    delete_selected_problems.short_description = _('Delete selected problems')
    
    def delete_model(self, request, obj):
        """Override to also delete the problem folder from filesystem."""
        import os
        import shutil
        from django.conf import settings
        
        # Store the code before deletion
        problem_code = obj.code
        
        # Delete from database (calls parent)
        super().delete_model(request, obj)
        
        # Delete problem folder from filesystem
        problems_dir = os.path.join(settings.BASE_DIR, 'problems')
        folder_path = os.path.join(problems_dir, problem_code)
        if os.path.exists(folder_path):
            try:
                shutil.rmtree(folder_path)
            except Exception as e:
                from django.contrib import messages
                messages.warning(request, f"Could not delete folder for '{problem_code}': {e}")

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
             path('bulk-upload/', self.bulk_upload_view, name='hpe_problem_bulk_upload'),
        ]
        return custom_urls + urls

    def bulk_upload_view(self, request):
        if request.method == 'POST':
            form = HPEProblemBulkUploadForm(request.POST, request.FILES)
            if form.is_valid():
                try:
                    uploaded_file = request.FILES['csv_file']
                    file_name = uploaded_file.name.lower()
                    
                    import csv
                    import io
                    import zipfile
                    import yaml
                    from django.core.files.base import ContentFile
                    from django.utils.text import slugify
                    from django.utils import timezone
                    from judge.models import ProblemData, Solution
                    
                    rows = []
                    
                    # Handle Excel files (.xlsx)
                    if file_name.endswith('.xlsx'):
                        from openpyxl import load_workbook
                        wb = load_workbook(filename=io.BytesIO(uploaded_file.read()))
                        ws = wb.active
                        for row in ws.iter_rows(values_only=True):
                            # Convert None to empty string and ensure all are strings
                            rows.append([str(cell) if cell is not None else '' for cell in row])
                    else:
                        # Handle CSV files
                        decoded_file = uploaded_file.read().decode('utf-8').splitlines()
                        reader = csv.reader(decoded_file)
                        rows = list(reader)
                    
                    # Expected format: Name, Body, Constraints, TL, ML, Difficulty, Solution, In1, Out1, In2, Out2...
                    
                    # Map difficulty names to group IDs
                    difficulty_map = {
                        'easy': 2,
                        'medium': 3,
                        'hard': 4,
                    }
                    
                    count = 0
                    success_results = []  # List of successfully created problems
                    error_results = []  # List of errors
                    
                    for row in rows:
                        if len(row) < 7: continue # Minimum fields (including difficulty and solution placeholder)
                        
                        # Skip rows where ALL columns are empty
                        if all(not cell or str(cell).strip() == '' for cell in row):
                            continue
                        
                        name = row[0].strip()
                        if not name: continue  # Skip rows with empty name
                        
                        body = row[1]
                        constraints = row[2]
                        
                        # Handle Null/empty time limit - default to 60 seconds (max reasonable)
                        tl_value = row[3].strip().lower() if row[3] else ''
                        if tl_value in ('null', 'none', '') or not tl_value:
                            time_limit = 60.0  # Default max time limit
                        else:
                            try:
                                time_limit = float(row[3])
                            except:
                                time_limit = 60.0
                        
                        # Handle Null/empty memory limit - default to 512MB (524288 KB)
                        ml_value = row[4].strip().lower() if row[4] else ''
                        if ml_value in ('null', 'none', '') or not ml_value:
                            memory_limit = 524288  # 512MB default
                        else:
                            try:
                                memory_limit = int(row[4])
                            except:
                                memory_limit = 524288
                        
                        # Handle Difficulty level - REQUIRED (easy/medium/hard)
                        # Get the difficulty value safely
                        raw_difficulty = row[5] if len(row) > 5 and row[5] else ''
                        difficulty_value = raw_difficulty.strip().lower()
                        
                        if difficulty_value not in difficulty_map:
                            error_results.append({'name': name, 'error': f"Invalid difficulty '{raw_difficulty}'. Must be easy, medium, or hard."})
                            continue  # Skip this row
                        group_id = difficulty_map[difficulty_value]
                        
                        # Handle Solution - Optional
                        solution_content = row[6].strip() if len(row) > 6 and row[6] else ''
                        
                        # Check for duplicate name in database
                        if Problem.objects.filter(name=name).exists():
                            error_results.append({'name': name, 'error': 'A problem with this name already exists.'})
                            continue
                        
                        # Check for duplicate body in database
                        if Problem.objects.filter(description__startswith=body[:500]).exists():
                            error_results.append({'name': name, 'error': 'A problem with similar body already exists.'})
                            continue
                        
                        # Generate unique problem code from name
                        base_code = slugify(name)[:20] or "prob"
                        code = base_code
                        suffix = 1
                        while Problem.objects.filter(code=code).exists():
                            code = f"{base_code}{suffix}"
                            suffix += 1
                        
                        # Parse test cases first to build Examples (now starting from index 7)
                        test_cases_data = row[7:]
                        examples_md = ""
                        example_num = 1
                        
                        # Build Examples section from test case pairs
                        for i in range(0, len(test_cases_data), 2):
                            if i + 1 >= len(test_cases_data): break
                            
                            in_data = test_cases_data[i]
                            out_data = test_cases_data[i+1]
                            
                            # Only show first 2 examples in description (visible to users)
                            if example_num <= 2:
                                examples_md += f"\n## Example {example_num}\n"
                                examples_md += f"**Input:**\n```\n{in_data}\n```\n\n"
                                examples_md += f"**Output:**\n```\n{out_data}\n```\n"
                            
                            example_num += 1
                        
                        # Build the full markdown description
                        description = body
                        
                        if examples_md:
                            description += f"\n{examples_md}"
                        
                        if constraints and constraints.lower() != 'none':
                            description += f"\n## Constraints\n{constraints}\n"
                            
                        # Create Problem with difficulty group
                        problem = Problem.objects.create(
                            code=code,
                            name=name,
                            description=description,
                            time_limit=time_limit,
                            memory_limit=memory_limit,
                            points=0,  # Default points (can be adjusted later)
                            is_public=False,
                            is_manually_managed=True,
                            group_id=group_id,  # Use mapped difficulty group
                            date=timezone.now()  # Set upload date to current time
                        )
                        
                        # M2M
                        if form.cleaned_data['creators']:
                            problem.authors.set(form.cleaned_data['creators'])
                        if form.cleaned_data['testers']:
                            problem.testers.set(form.cleaned_data['testers'])
                        if form.cleaned_data['allowed_languages']:
                            problem.allowed_languages.set(form.cleaned_data['allowed_languages'])
                            
                        # Create Solution if provided
                        if solution_content:
                            Solution.objects.create(
                                problem=problem,
                                content=solution_content,
                                is_public=False,
                                publish_on=timezone.now()
                            )
                            
                        # Test Cases - Create folder structure in site/problems/{code}/
                        if test_cases_data:
                            import os
                            from django.conf import settings
                            
                            # Get the problems directory path
                            problems_dir = os.path.join(settings.BASE_DIR, 'problems')
                            problem_dir = os.path.join(problems_dir, code)
                            
                            # Create the problem directory
                            os.makedirs(problem_dir, exist_ok=True)
                            
                            # Build init.yml content
                            init_yml_content = {
                                'name': name,
                                'code': code,
                                'type': 'standard',
                                'validator': 'token',
                                'limits': {
                                    'time': time_limit,
                                    'memory': memory_limit
                                },
                                'archive': 'testcases.zip',
                                'cases': []
                            }
                            
                            # Create individual test case files and add to zip
                            zip_buffer = io.BytesIO()
                            
                            with zipfile.ZipFile(zip_buffer, 'w') as zf:
                                case_idx = 1
                                for i in range(0, len(test_cases_data), 2):
                                    if i + 1 >= len(test_cases_data): break
                                    
                                    in_data = test_cases_data[i]
                                    out_data = test_cases_data[i+1]
                                    
                                    in_name = f"{case_idx}.in"
                                    out_name = f"{case_idx}.out"
                                    
                                    # Write individual files to folder
                                    with open(os.path.join(problem_dir, in_name), 'w') as f:
                                        f.write(in_data)
                                    with open(os.path.join(problem_dir, out_name), 'w') as f:
                                        f.write(out_data)
                                    
                                    # Add to zip
                                    zf.writestr(in_name, in_data)
                                    zf.writestr(out_name, out_data)
                                    
                                    init_yml_content['cases'].append({
                                        'in': in_name,
                                        'out': out_name
                                    })
                                    case_idx += 1
                            
                            if init_yml_content['cases']:
                                # Write init.yml to folder with specific field order
                                yml_lines = [
                                    f"name: {name}",
                                    f"code: {code}",
                                    "type: standard",
                                    "validator: token",
                                    "limits:",
                                    f"  time: {time_limit}",
                                    f"  memory: {memory_limit}",
                                    "archive: testcases.zip",
                                    "cases:",
                                ]
                                for case in init_yml_content['cases']:
                                    yml_lines.append(f"  - in: {case['in']}")
                                    yml_lines.append(f"    out: {case['out']}")
                                
                                with open(os.path.join(problem_dir, 'init.yml'), 'w') as f:
                                    f.write('\n'.join(yml_lines) + '\n')
                                
                                # Write testcases.zip to folder
                                with open(os.path.join(problem_dir, 'testcases.zip'), 'wb') as f:
                                    f.write(zip_buffer.getvalue())
                        
                        count += 1
                        success_results.append({'name': name, 'code': code})
                    
                    # Stay on same page and show results
                    return render(request, 'hpe_admin/bulk_problem_upload.html', {
                        'form': HPEProblemBulkUploadForm(),  # Fresh form
                        'title': _('Bulk Upload Problems'),
                        'upload_complete': True,
                        'success_results': success_results,
                        'error_results': error_results,
                        'total_success': count,
                        'total_errors': len(error_results)
                    })
                    
                except Exception as e:
                    error_results = [{'name': 'File Error', 'error': str(e)}]
                    return render(request, 'hpe_admin/bulk_problem_upload.html', {
                        'form': HPEProblemBulkUploadForm(),
                        'title': _('Bulk Upload Problems'),
                        'upload_complete': True,
                        'success_results': [],
                        'error_results': error_results,
                        'total_success': 0,
                        'total_errors': 1
                    })
        else:
            form = HPEProblemBulkUploadForm()
            
        return render(request, 'hpe_admin/bulk_problem_upload.html', {
            'form': form,
            'title': _('Bulk Upload Problems')
        })
    
    def get_form(self, request, obj=None, **kwargs):
        # Bypass ProblemAdmin.get_form
        return super(ProblemAdmin, self).get_form(request, obj, **kwargs)
    # Note: Bulk Upload button will be added via template

class HPEMCQQuestionAdmin(MCQQuestionAdmin):
    form = HPEMCQForm
    change_list_template = 'hpe_admin/mcq_change_list.html'
    
    # Simplified list view (difficulty and types shown only in edit form)
    list_display = ['code', 'question_type', 'points', 'is_public', 'date']
    list_filter = ['group', 'types', 'question_type', 'is_public']  # Filter by difficulty, types, question type, and visibility
    search_fields = ['code', 'description']
    filter_horizontal = ['types']  # Use horizontal filter widget for types
    
    def difficulty_display(self, obj):
        """Display group as 'Difficulty'"""
        if obj.group:
            return obj.group.full_name
        return '-'
    difficulty_display.short_description = 'Difficulty'
    
    def types_display(self, obj):
        """Display types as comma-separated list"""
        types = obj.types.all()
        if types:
            return ', '.join([t.full_name for t in types])
        return '-'
    types_display.short_description = 'Types'
    
    def get_form(self, request, obj=None, **kwargs):
        # Bypass MCQQuestionAdmin.get_form
        return super(MCQQuestionAdmin, self).get_form(request, obj, **kwargs)
    
    # Fieldsets with Difficulty and Types
    fieldsets = (
        (None, {'fields': ('code', 'authors', 'description')}),
        (_('Categorization'), {'fields': ('group', 'types')}),
        (_('Settings'), {'fields': ('question_type', 'points', 'partial_credit', 'explanation')}),
    )
    
    def get_actions(self, request):
        """Re-enable delete action with custom implementation to handle distinct()"""
        actions = super().get_actions(request)
        # Add custom delete action that works with distinct querysets
        actions['delete_selected'] = (self.delete_selected_mcqs, 'delete_selected', _('Delete selected MCQs'))
        return actions
    
    def delete_selected_mcqs(self, modeladmin, request, queryset):
        """Custom delete action with confirmation page that handles distinct querysets"""
        from django.contrib import messages
        from django.template.response import TemplateResponse
        
        # Get the IDs first to avoid the distinct() issue
        mcq_ids = list(queryset.values_list('id', flat=True))
        mcqs_to_delete = MCQQuestion.objects.filter(id__in=mcq_ids)
        count = len(mcq_ids)
        
        if count == 0:
            messages.warning(request, "No MCQs selected.")
            return None
        
        # Check if user confirmed deletion
        if request.POST.get('post') == 'yes':
            # Delete from database
            mcqs_to_delete.delete()
            messages.success(request, f"Successfully deleted {count} MCQ{'' if count == 1 else 's'}.")
            return None
        
        # Show confirmation page
        context = {
            **self.admin_site.each_context(request),
            'title': _('Are you sure?'),
            'mcqs': mcqs_to_delete,
            'count': count,
            'action_checkbox_name': '_selected_action',
            'queryset': queryset,
            'opts': self.model._meta,
            'media': self.media,
        }
        
        return TemplateResponse(request, 'hpe_admin/confirm_delete_mcqs.html', context)
    delete_selected_mcqs.short_description = _('Delete selected MCQs')
    
    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path('bulk-upload/', self.bulk_upload_view, name='hpe_mcq_bulk_upload'),
        ]
        return custom_urls + urls
    
    def bulk_upload_view(self, request):
        from judge.models import MCQQuestion, MCQOption
        from django.utils.text import slugify
        import csv
        import io
        
        if request.method == 'POST':
            form = HPEMCQBulkUploadForm(request.POST, request.FILES)
            if form.is_valid():
                try:
                    uploaded_file = request.FILES['csv_file']
                    file_name = uploaded_file.name.lower()
                    
                    rows = []
                    
                    # Handle Excel files (.xlsx)
                    if file_name.endswith('.xlsx'):
                        from openpyxl import load_workbook
                        wb = load_workbook(filename=io.BytesIO(uploaded_file.read()))
                        ws = wb.active
                        for row in ws.iter_rows(values_only=True):
                            rows.append([str(cell) if cell is not None else '' for cell in row])
                    else:
                        # Handle CSV files
                        decoded_file = uploaded_file.read().decode('utf-8').splitlines()
                        reader = csv.reader(decoded_file)
                        rows = list(reader)
                    
                    # Format: Question, Option1, Option2, Option3, Option4, Answer(s)...
                    count = 0
                    success_results = []  # List of successfully created MCQs
                    error_results = []  # List of errors
                    
                    for row in rows:
                        if len(row) < 6: continue  # Minimum: Question + 4 options + 1 answer
                        
                        # Skip rows where ALL columns are empty
                        if all(not cell or str(cell).strip() == '' for cell in row):
                            continue
                        
                        question_text = row[0].strip()
                        if not question_text: continue  # Skip rows with empty question
                        
                        # Check for duplicate question text
                        if MCQQuestion.objects.filter(description=question_text).exists():
                            error_results.append({'name': question_text[:50] + '...', 'error': 'This question already exists.'})
                            continue
                        
                        # Get the 4 options
                        options = [row[i].strip() if i < len(row) else '' for i in range(1, 5)]
                        
                        # Get answer columns (column 6 onwards) - these are the ANSWER TEXTS
                        answers_raw = [row[i].strip().lower() for i in range(5, len(row)) if i < len(row) and row[i] and row[i].strip()]
                        
                        # Match answers to options by TEXT (case-insensitive)
                        correct_options = set()
                        for idx, option in enumerate(options):
                            if option.lower() in answers_raw:
                                correct_options.add(idx + 1)  # 1-indexed
                        
                        if not correct_options:
                            error_results.append({'name': question_text[:50] + '...', 'error': 'No matching answers found. Check answer text matches option text.'})
                            continue
                        
                        # Determine question type
                        question_type = 'MULTIPLE' if len(correct_options) > 1 else 'SINGLE'
                        
                        # Generate unique code
                        base_code = slugify(question_text[:15]).replace('-', '') or "mcq"
                        code = base_code[:20]
                        suffix = 1
                        while MCQQuestion.objects.filter(code=code).exists():
                            code = f"{base_code[:17]}{suffix}"
                            suffix += 1
                        
                        # Create MCQ Question
                        from django.utils import timezone
                        from judge.models.problem import ProblemGroup
                        
                        # Get or create 'uncategorized' difficulty group
                        uncategorized_group, _created = ProblemGroup.objects.get_or_create(
                            name='uncategorized',
                            defaults={'full_name': 'Uncategorized'}
                        )
                        
                        mcq = MCQQuestion.objects.create(
                            code=code,
                            description=question_text,  # Full question text
                            question_type=question_type,
                            points=1.0,
                            is_public=False,
                            date=timezone.now(),  # Set published date to current time
                            group=uncategorized_group  # Default to uncategorized difficulty
                        )
                        
                        # Set creators if specified
                        if form.cleaned_data['creators']:
                            mcq.authors.set(form.cleaned_data['creators'])
                        
                        # Create options
                        for idx, option_text in enumerate(options):
                            if option_text:  # Only create non-empty options
                                MCQOption.objects.create(
                                    question=mcq,
                                    option_text=option_text,
                                    is_correct=(idx + 1) in correct_options,
                                    order=idx
                                )
                        
                        count += 1
                        success_results.append({'name': question_text[:50] + '...', 'code': code})
                    
                    # Stay on same page and show results
                    return render(request, 'hpe_admin/bulk_mcq_upload.html', {
                        'form': HPEMCQBulkUploadForm(),  # Fresh form
                        'title': _('Bulk Upload MCQ Questions'),
                        'upload_complete': True,
                        'success_results': success_results,
                        'error_results': error_results,
                        'total_success': count,
                        'total_errors': len(error_results)
                    })
                    
                except Exception as e:
                    error_results = [{'name': 'File Error', 'error': str(e)}]
                    return render(request, 'hpe_admin/bulk_mcq_upload.html', {
                        'form': HPEMCQBulkUploadForm(),
                        'title': _('Bulk Upload MCQ Questions'),
                        'upload_complete': True,
                        'success_results': [],
                        'error_results': error_results,
                        'total_success': 0,
                        'total_errors': 1
                    })
        else:
            form = HPEMCQBulkUploadForm()
            
        return render(request, 'hpe_admin/bulk_mcq_upload.html', {
            'form': form,
            'title': _('Bulk Upload MCQ Questions')
        })
    # Note: Bulk Upload button will be added via template


# Register models
hpe_admin_site.register(Problem, HPEProblemAdmin)
hpe_admin_site.register(MCQQuestion, HPEMCQQuestionAdmin)
hpe_admin_site.register(Contest, HPEContestAdmin)
